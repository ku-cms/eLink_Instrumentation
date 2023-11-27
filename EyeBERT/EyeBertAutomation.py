# EyeBertAutomation.py
#
# Test script for using pyautogui to control the Vivado EyeBert
# program for CMS twisted pair BERT testing.
#
# Automatically generate eye diagram, save screen shot of diagram
# and extract the "Open Area" value from the measurement
#
# To Do : Add appending to XLS file with various cable info
#       : Get cable specifics from operator
#       : control EyeBERT relay board
#       : repeat tests as needed
#       : much better error recovery & data validation!

version = 1.07

from colorama import Fore, Back, Style, init

init(convert=True)
print(Fore.GREEN + "KU-CMS KC705 EyeBERT Automated Test")
print(f"Version {version}")
print(Fore.RESET + "Loading libraries...")

import pyautogui as pygui
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
import os
import shutil
import time
import sys
import ctypes
from pathvalidate import is_valid_filename
import eyebertserial

#
# The e-link connection mapping is defined by a dictionary
# - Define e-link mappings as needed
# - Use the desired e-link mapping
#

# SMA cables: testing in loopback mode
mapping_SMA_test = {
    "name" : "garbage",
    "cmd" : {"tx" : "0", "rx" : "0"},
    "d0"  : {"tx" : "3", "rx" : "3"}
}

# Type 1 e-links
mapping_type1 = {
    "name" : "type 1 tfpix",
    "cmd" : {"tx" : "7", "rx" : "7"},
    "d0"  : {"tx" : "0", "rx" : "0"},
    "d1"  : {"tx" : "1", "rx" : "1"},
    "d2"  : {"tx" : "2", "rx" : "2"},
    "d3"  : {"tx" : "3", "rx" : "3"}
}

# Type 5 e-links
mapping_type5 = {
    "name" : "type 5 tbpix",
    "cmd" : {"tx" : "7", "rx" : "7"},
    "d0"  : {"tx" : "0", "rx" : "0"},
    "d1"  : {"tx" : "1", "rx" : "1"},
    "d2"  : {"tx" : "2", "rx" : "2"},
    "d3"  : {"tx" : "3", "rx" : "3"}
}

# Choose e-link mapping:
cable_mapping = mapping_type5
print(Fore.GREEN + "Mapping Dictionary : " + Fore.RED + 
      cable_mapping["name"] + Fore.GREEN + " selected.")

test_results = {}

pygui.PAUSE = 0.5

#
# open serial control of relay board
#
eb = eyebertserial.EyeBERTRelayControl()
if eb.initialize() == False :
    print(Fore.RED + "Terminating code 1")
    sys.exit(1)
eb.Blinky(5)

#
# get operator name/initials
#
operator = input(Fore.RED + "Enter operator name: " + Fore.GREEN)
operator = operator.strip()
while operator == "" :
    operator = input(Fore.RED + "Enter operator name: " + Fore.GREEN)
    operator = operator.strip()

#operator = operator.upper()

#
# get a valid cable name to use as directory and root filename
#
file_path = "C:/Users/Public/Documents/automation_results"
r_file_path = "R:/BEAN_GRP/EyeBERTAutomation/automation_results"
is_valid = False
while is_valid == False :
    filename = input(Fore.RED + "Enter cable name: " + Fore.GREEN)
    is_valid = is_valid_filename(filename)
    if is_valid == False :
        print(Fore.RED + f"{filename} cannot be used as a directory or filename. Re-enter.")

#
# get ready to use this as our destination path
#
cable_path = file_path + "/" + filename
r_cable_path = r_file_path + "/" + filename
isExist = os.path.exists(cable_path)
if isExist == False :
    # create the path
    os.makedirs(cable_path)

isExist = os.path.exists(r_cable_path)
if isExist == False :
    #create the path
    os.makedirs(r_cable_path)
    
#
# get serial number of test boards
#
left_serialnumber = input(Fore.RED + "Enter SN of left board: " + Fore.GREEN)
left_serialnumber = left_serialnumber.strip()
while left_serialnumber == "" :
    left_serialnumber = input(Fore.RED + "Enter SN of left board: " + Fore.GREEN)
    left_serialnumber = left_serialnumber.strip()

right_serialnumber = input(Fore.RED + "Enter SN of right board: " + Fore.GREEN)
right_serialnumber = right_serialnumber.strip()
while right_serialnumber == "" :
    right_serialnumber = input(Fore.RED + "Enter SN of right board: " + Fore.GREEN)
    right_serialnumber = right_serialnumber.strip()
    
left_serialnumber = left_serialnumber.upper()
right_serialnumber = right_serialnumber.upper()
    
#
# brief operator notes
#
operator_notes = input(Fore.RED + "Operator notes: " + Fore.GREEN)


#
# guarantee caps lock is off
#
if ctypes.WinDLL("User32.dll").GetKeyState(0x14) :
    print(Fore.RED + "CAPSLOCK is on - turning off." + Fore.RESET)
    pygui.press('capslock')

# loop starts here

keys = list(cable_mapping.keys())
for key in keys :
    if key == "name" :
        continue
    temp_name = filename + "_" + str(key) # change for looping version
    test_name = temp_name.replace(" ", "_")
    txpath = b"tx " + bytes(cable_mapping[key]['tx'], 'utf-8')
    rxpath = b"rx " + bytes(cable_mapping[key]['rx'], 'utf-8')
    
    print(Fore.GREEN + Style.BRIGHT + f"{test_name}" + Style.NORMAL)
    print(Fore.GREEN + f"\ttxpath = {txpath.decode()}")
    print(Fore.GREEN + f"\trxpath = {rxpath.decode()}")

    eb.connection(txpath+b"\r\n")
    eb.connection(rxpath+b"\r\n")
    eb.LED(2,"ON")

    # if old temp.csv exists, erase
    if os.path.exists(file_path+"/temp.csv") :
        os.remove(file_path+"/temp.csv")

    # makes assumption that EyeBERT KC704 (Vivado) is already running
    # with bitstream loaded
    # guarantee it is maximized and has focus
    try :
        pygui.getWindowsWithTitle("Vivado 2020.2")[0].maximize()
        time.sleep(0.5)
        pygui.getWindowsWithTitle("Vivado 2020.2")[0].activate()
        time.sleep(0.5)
        # send F5 to reset layout of Vivado
        pygui.press('f5')
    except :
        print(Fore.RED + "Vivado 2020.2 and EyeBert bitstream not running")
        print(Fore.RED + "Terminating code 2" + Fore.RESET)
        sys.exit(2)

    # create custom "source eye_and_save.tcl" with the cable name & path included
    eye_and_save = "C:/Users/Public/Documents/cable_tests/eye_and_save.tcl"
    with open (eye_and_save, 'w') as f :
        f.write("source create_scan_0.tcl\r\n")
        f.write(f"set_property DESCRIPTION {test_name} [get_hw_sio_scans SCAN_0]\r\n")
        f.write("run_hw_sio_scan [lindex [get_hw_sio_scans {SCAN_0}] 0]\r\n")
        f.write("wait_on_hw_sio_scan [lindex [get_hw_sio_scans {SCAN_0}] 0]\r\n")
        f.write('write_hw_sio_scan -force "C:/Users/Public/Documents/automation_results/temp.csv" [get_hw_sio_scans {SCAN_0}]\r\n')

    # send tcl "source eye_and_save.tcl"
    result = pygui.locateCenterOnScreen('tcl_console.png', grayscale=True)
    if result == None :
        result = pygui.locateCenterOnScreen('light_tcl_console.png', grayscale=True)

    if result == None :
        # handle error
        print(Fore.RED + "Unable to locate TCL Console tab.")
        sys.exit(9)

    pygui.click(result) # bring tab to focus
    time.sleep(0.25)
    pygui.press('tab') # move focus to text entry box of tab
    time.sleep(0.25)

    # launch the test via TCL script
    pygui.write("source eye_and_save.tcl\n", interval = 0.01)

    # wait a bit
    print(Fore.GREEN + "Pausing to allow EyeBERT to complete...")
    time.sleep(15)

    # loop looking for temp.csv
    print(Fore.GREEN + "Waiting for data file from Vivado...")
    time_to_wait = 20
    time_counter = 0
    while not os.path.exists(file_path+"/temp.csv") :
        time.sleep(1)
        time_counter += 1
        if time_counter > time_to_wait : break

    eb.LED(2,"OFF")

    # get data
    print(Fore.GREEN + "Parsing CSV file for test results...")
    open_area = -999.9
    with open(file_path+"/temp.csv", 'r') as file :
        csvreader = csv.reader(file, delimiter=',')
        # find the Open Area
        for row in csvreader :
            if row[0] == "Open Area" :
                open_area = int(row[1])
                break

    with open(file_path+"/temp.csv", 'r') as file :
        search = list(csv.reader(file)) # convert to a list we can index through
        # find the top of eye
        top_of_eye = -999.9 # nonsense values
        # use column AH, start at row 22
        for r in range(21, 45, 1) :
            value = float(search[r][33])
            if value < 2.0e-7 : 
                # we have found it
                top_of_eye = float(search[r-1][0])
                break
        # find the bottom of eye
        bottom_of_eye = 999.9
        for r in range(44, 20, -1) :
            value = float(search[r][33])
            if value < 2.0e-7 :
                # we found it!
                bottom_of_eye = float(search[r+1][0])
                break

    print(Fore.GREEN + Style.BRIGHT + f"Open area = {open_area}")
    print(Fore.GREEN + f"top_of_eye = {top_of_eye}")
    print(Fore.GREEN + f"bottom_of_eye = {bottom_of_eye}" + Style.NORMAL)

    src = file_path+"/temp.csv"
    dest = cable_path + "/" + test_name + ".csv"
    # check if destionation already exists
    counter = 2
    while os.path.exists(dest) == True :
        dest = cable_path + "/" + test_name + "_" + str(counter) + ".csv"
        counter = counter + 1

    os.rename(src,dest)

    copyfilename = dest.replace(file_path+"/","")
    newdest = r_file_path + "/" + copyfilename
    shutil.copyfile(dest, newdest)
    
    # make the screen capture
    print(Fore.GREEN + "Creating screen capture...")
    pygui.moveTo(355,800)
    pygui.click()
    screenshot = cable_path + "/" + test_name + ".png"
    counter = 2
    while os.path.exists(screenshot) == True :
        screenshot = cable_path + "/" + test_name + "_" + str(counter) + ".png"
        counter = counter + 1

    pygui.screenshot(screenshot,(360,190,1300,540))
    copyfilename = screenshot.replace(file_path+"/","")
    newdest = r_file_path + "/" + copyfilename
    shutil.copyfile(screenshot, newdest)
    
    # add results to dataset for future write
    now = datetime.datetime.now()
    channel_name = key
    test_results.update(
        {key : 
         {"test_name" : test_name.replace("_"+channel_name,""),
          "channel" : channel_name,
          "date" : now.strftime("%Y-%m-%d"),
          "time" : now.strftime("%H:%M:%S"),
          "open_area" : open_area, 
          "top_eye" : top_of_eye, 
          "bottom_eye" : bottom_of_eye,
          "operator" : operator,
          "left_SN" : left_serialnumber, 
          "right_SN" : right_serialnumber,
          "notes" : operator_notes}
        }
    )
#end keys loop


# update XLS file, create new entries as needed
wb = Workbook()
file_name = "EyeBERTautomation.xlsx"
path = "R:/BEAN_GRP/EyeBertAutomation/"
keep_trying = True
while keep_trying :
    try :
        os.rename(path + file_name, path + file_name)
        keep_trying = False
    except OSError:
        pygui.getWindowsWithTitle("Vivado 2020.2")[0].minimize()
        print(Fore.RED + "EyeBERTautomation.xlxs summary file is open. Please close")
        x=input(Fore.RED + "Press ENTER when ready to retry" + Fore.GREEN)
        keep_trying = True

isExist = os.path.exists(path + file_name)
if isExist == False :
    # create file
    print(Fore.LIGHTRED_EX + "\tXLSX summary file does not exist. Creating file.")
    ws = wb.active
    newdata = ["Cable name", "channel", "date", "time", "open_area", "top_eye",
               "bottom_eye", "operator", "left_SN", "right_SN", "notes"]
    ws.append(newdata)
    col = get_column_letter(1)
    ws.column_dimensions[col].bestFit = True
    wb.save(path + file_name)
else :
    # load existing copy
    print(Fore.GREEN + "Opening XLSX summary file")
    wb = load_workbook(filename = path + file_name)
    ws = wb.active

print(Fore.GREEN + "Adding data...")
cablename = test_name
keys = list(test_results)
for key in keys :
    newdata = [test_results[key]["test_name"],
               test_results[key]["channel"],
               test_results[key]["date"],
               test_results[key]["time"],
               test_results[key]["open_area"],
               test_results[key]["top_eye"],
               test_results[key]["bottom_eye"],
               test_results[key]["operator"],
               test_results[key]["left_SN"],
               test_results[key]["right_SN"],
               test_results[key]["notes"]]
    ws.append(newdata)
col = get_column_letter(1)
ws.column_dimensions[col].bestFit = True
wb.save(path + file_name)

pygui.getWindowsWithTitle("Vivado 2020.2")[0].minimize()
print(Fore.GREEN + Style.BRIGHT + "Done!" + Fore.RESET + Style.RESET_ALL)

#excel_start_return = os.system('start "excel" ' + path + file_name)


    
