# EyeBertAutomationRevB.py
#
# Developed by the KU CMS group.
#
# Authors:
# - Rob Young
# - Nora Manolescu
# - Caleb Smith
#
# Used for the following measurements for twisted pair electrical links (e-links):
#
# 1. Performs 4-point DC resistance measurements of all e-link channels (both positive and negative wires)
#    using a Keithley, relay boards, SMA cables, and adapter boards.
#    Saves 4-point DC resistance measurement results.
#    Includes functionality for calibrating 4-point DC resistance for every e-link channel.
#
# 2. Performs "Eye BERT" area measurements (including eye-diagram template analysis) of all e-link channels (each differential pair)
#    using pyautogui to control the Vivado "Eye BERT" program, where the hardware setup is a KC705, relay boards, SMA cables, and adapter boards.
#    Saves eye-diagram and template analysis plots and records open area, height, and template analysis results.
#
# To Do
#   : Get parameters (which tests to run) from operator
#   : For 4-point DC calibration, automatically create new file name (default) or let user overwrite existing file
#   : For each 4-point DC calibration measurement, allow user to accept value or redo calibration measurement 
#   : repeat tests as needed
#   : much better error recovery & data validation!

# Hardware requirements:
# - KC705
# - Digital Multimeter (DMM)
# - Relay board (Rev B)
# - SMA adapter boards: one 45-pin Kyocera board and three 15-pin Molex boards 
# - SMA cables

# Software requirements:
# - Computer: Windows 10 PC 
# - Languages: Python 3, TCL, Windows Batch Script 
# - Applications: Vivado 2020.2 
# - Code Repository: https://github.com/ku-cms/eLink_Instrumentation 

# version
version = 2.3

from template_analysis_windows_RevB import EyeBERTFile, Reference
from colorama import Fore, Back, Style, init

init(convert=True)
print(Fore.GREEN + "KU-CMS 4-point DC Resistance and KC705 EyeBERT Automated Test")
print(f"Version {version:.2f}")
print(Fore.RESET + "Loading libraries...")

import pyautogui as pygui
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
import os
import shutil
import time
import sys
import ctypes
from pathvalidate import is_valid_filename
import eyebertserial
import dmmserial
import tools
import json

# get 4-point DC calibration file based on cable type: only used for Rev B relay board
def GetDCCalibrationFile(cable_type):
    if cable_type in ["1", "5K", "5K2"]:
        return "4_point_DC_Calibration_v1.json"
    elif cable_type in ["3p2", "2p2"]:
        return "4_point_DC_CalibrationRevB_Type3p2_v2.json"
    elif cable_type in ["2p3", "1p3"]:
        return "4_point_DC_CalibrationRevB_Type2p3_v2.json"
    else:
        print(Fore.RED + f"ERROR: There is no calibration file available for cable type {cable_type}." + Fore.GREEN)
        return ""
    
def main():
    # parameters
    # TODO: let user specify parameters for what test(s) to run
    verbose                     = False
    RUN_4PT_DC_RES_CALIBRATION  = False
    RUN_4PT_DC_RES              = True
    RUN_EYE_BERT_AREA           = True
    pygui.PAUSE = 0.5
    
    # dictionaries to save results
    dc_resistance_results = {}
    eye_bert_results = {}

    #
    # The e-link connection mapping is defined by a dictionary
    # - Define e-link mappings as needed
    # - Use the desired e-link mapping
    #

    # SMA cables: testing in loopback mode
    mapping_SMA_test = {
        "name"  : "example",
        "cmd"   : {"tx" : "0", "rx" : "0"},
        "d0"    : {"tx" : "3", "rx" : "3"}
    }

    # SMAs need to be connected with the correct channel mapping (based on e-link mapping).
    # SMAs need to be connected with the correct polarity (based on e-link mapping).

    # Type 1 e-links
    mapping_type1 = {
        "name"  : "Type 1",
        "cmd"   : {"tx" : "7", "rx" : "7"},
        "d0"    : {"tx" : "0", "rx" : "0"},
        "d1"    : {"tx" : "1", "rx" : "1"},
        "d2"    : {"tx" : "2", "rx" : "2"},
        "d3"    : {"tx" : "3", "rx" : "3"}
    }

    # TBPX Type 5 e-links (works for 5K and 5K2)
    # Note: Type 5K2 has inverted polarity for CMD and D2 compared to type 5K.
    #       For Type 5K2, you need to account for this by swapping P/N SMA cables for CMD and D2.
    mapping_type5 = {
        "name"  : "TBPX Type 5",
        "cmd"   : {"tx" : "5", "rx" : "9"},
        "d0"    : {"tx" : "1", "rx" : "11"},
        "d1"    : {"tx" : "2", "rx" : "10"},
        "d2"    : {"tx" : "3", "rx" : "4"},
        "d3"    : {"tx" : "4", "rx" : "6"}
    }

    # -------------------------------------------- #
    # e-link mappings v2 for Rev B relay board
    # use branch and channel in key
    # -------------------------------------------- #

    # TFPX Type 3.2 e-links
    mapping_type3p2 = {
        "name"  : "TFPX Type 3p2",
        "A_cmd"   : {"tx" : "5",    "rx" : "3"},
        "A_d0"    : {"tx" : "1",    "rx" : "4"},
        "A_d2"    : {"tx" : "3",    "rx" : "2"},
        "B_cmd"   : {"tx" : "10",   "rx" : "7"},
        "B_d0"    : {"tx" : "6",    "rx" : "8"},
        "B_d2"    : {"tx" : "8",    "rx" : "6"},
        "C_cmd"   : {"tx" : "15",   "rx" : "9"},
        "C_d0"    : {"tx" : "11",   "rx" : "11"},
        "C_d2"    : {"tx" : "13",   "rx" : "10"}
    }

    # TFPX Type 2.2 e-links
    mapping_type2p2 = {
        "name"  : "TFPX Type 2p2",
        "A_cmd"   : {"tx" : "5",    "rx" : "3"},
        "A_d0"    : {"tx" : "1",    "rx" : "4"},
        "A_d2"    : {"tx" : "3",    "rx" : "2"},
        "C_cmd"   : {"tx" : "15",   "rx" : "9"},
        "C_d0"    : {"tx" : "11",   "rx" : "11"},
        "C_d2"    : {"tx" : "13",   "rx" : "10"}
    }

    # TFPX Type 2.3 e-links
    # Note: Channel labels are different than types 3.2 and 2.2!
    mapping_type2p3 = {
        "name"  : "TFPX Type 2p3",
        "A_cmd"   : {"tx" : "5",    "rx" : "1"},
        "A_d2"    : {"tx" : "2",    "rx" : "6"},
        "A_d1"    : {"tx" : "3",    "rx" : "4"},
        "A_d0"    : {"tx" : "4",    "rx" : "2"},
        "B_cmd"   : {"tx" : "10",   "rx" : "9"},
        "B_d2"    : {"tx" : "7",    "rx" : "11"},
        "B_d1"    : {"tx" : "8",    "rx" : "10"},
        "B_d0"    : {"tx" : "9",    "rx" : "8"}
    }

    # TFPX Type 1.3 e-links
    # Note: Channel labels are different than types 3.2 and 2.2!
    mapping_type1p3 = {
        "name"  : "TFPX Type 1p3",
        "A_cmd"   : {"tx" : "5",    "rx" : "1"},
        "A_d2"    : {"tx" : "2",    "rx" : "6"},
        "A_d1"    : {"tx" : "3",    "rx" : "4"},
        "A_d0"    : {"tx" : "4",    "rx" : "2"}
    }

    # cable mappings for supported e-link types
    cable_mappings = {
        "1"     : mapping_type1,
        "5K"    : mapping_type5,
        "5K2"   : mapping_type5,
        "3p2"   : mapping_type3p2,
        "2p2"   : mapping_type2p2,
        "2p3"   : mapping_type2p3,
        "1p3"   : mapping_type1p3
    }
    
    # channels for Rev B relay board for supported e-link types
    # - for e-links without branches, entries should include channel and sign (e.g. cmd_p)
    # - for e-links with branches, entries should include branch, channel, and sign (e.g. A_cmd_p)
    cable_channels = {
        "1"     : ["cmd_p", "cmd_n", "d0_p", "d0_n", "d1_p", "d1_n", "d2_p", "d2_n", "d3_p", "d3_n"],
        "5K"    : ["cmd_p", "cmd_n", "d0_p", "d0_n", "d1_p", "d1_n", "d2_p", "d2_n", "d3_p", "d3_n"],
        "5K2"   : ["cmd_p", "cmd_n", "d0_p", "d0_n", "d1_p", "d1_n", "d2_p", "d2_n", "d3_p", "d3_n"],
        "3p2"   : ["A_cmd_p", "A_cmd_n", "A_d0_p", "A_d0_n", "A_d2_p", "A_d2_n", "B_cmd_p", "B_cmd_n", "B_d0_p", "B_d0_n", "B_d2_p", "B_d2_n", "C_cmd_p", "C_cmd_n", "C_d0_p", "C_d0_n", "C_d2_p", "C_d2_n"],
        "2p2"   : ["A_cmd_p", "A_cmd_n", "A_d0_p", "A_d0_n", "A_d2_p", "A_d2_n", "C_cmd_p", "C_cmd_n", "C_d0_p", "C_d0_n", "C_d2_p", "C_d2_n"],
        "2p3"   : ["A_cmd_p", "A_cmd_n", "A_d0_p", "A_d0_n", "A_d1_p", "A_d1_n", "A_d2_p", "A_d2_n", "B_cmd_p", "B_cmd_n", "B_d0_p", "B_d0_n", "B_d1_p", "B_d1_n", "B_d2_p", "B_d2_n"],
        "1p3"   : ["A_cmd_p", "A_cmd_n", "A_d0_p", "A_d0_n", "A_d1_p", "A_d1_n", "A_d2_p", "A_d2_n"]
    }

    # cable branches based on cable type
    cable_branches = {
        "3p2" : ["A", "B", "C"],
        "2p2" : ["A", "C"],
        "2p3" : ["A", "B"],
        "1p3" : ["A"]
    }

    # supported cable types
    cable_types = list(cable_mappings.keys())

    #
    # open serial control of relay board
    #
    eb = eyebertserial.EyeBERTRelayControl()
    if eb.initialize() == False :
        print(Fore.RED + "Terminating code 1 eyebert relay init")
        sys.exit(1)
    eb.Blinky(5)

    #
    # open dmm
    # 
    dmm = None
    if RUN_4PT_DC_RES or RUN_4PT_DC_RES_CALIBRATION:
        dmm = dmmserial.Keithley2000DMMControl()
        if dmm.initialize() == False :
            print(Fore.RED + "Terminating code 1 DMM init")
            sys.exit(1)

    #
    # get operator name/initials
    #
    operator = input(Fore.RED + "Enter operator name: " + Fore.GREEN)
    operator = operator.strip()
    while operator == "" :
        operator = input(Fore.RED + "Enter operator name: " + Fore.GREEN)
        operator = operator.strip()

    #operator = operator.upper()

    #
    # get a valid cable name to use as directory and root filename
    #
    file_path = "C:/Users/Public/Documents/automation_results"
    r_file_path = "R:/BEAN_GRP/EyeBERTAutomation/automation_results"
    is_valid = False
    while is_valid == False:
        filename = input(Fore.RED + "Enter cable number: " + Fore.GREEN)
        is_valid = is_valid_filename(filename)
        if is_valid == False:
            print(Fore.RED + f"{filename} cannot be used as a directory or filename. Re-enter.")

    # Assume that the filename is the cable number
    cable = filename

    #
    # get cable type
    #
    is_valid = False
    while is_valid == False:
        cable_type = input(Fore.RED + f"Enter cable type {cable_types}: " + Fore.GREEN)
        is_valid = tools.is_valid_cable_type(cable_types, cable_type)
        if is_valid == False:
            print(Fore.RED + f"{cable_type} is not a valid cable type. Re-enter a valid cable type: {cable_types}.")
    
    # assign mapping and channels based on cable type
    cable_mapping   = cable_mappings[cable_type]
    all_channels    = cable_channels[cable_type]
    print(Fore.GREEN + "Mapping Dictionary: " + Fore.RED + cable_mapping["name"] + Fore.GREEN + " selected.")

    #
    # get cable branch
    #
    # For the Rev B relay board, the user does not need to enter a branch!
    branches = []
    branch = ""

    #
    # get ready to use this as our destination path
    #
    cable_path = file_path + "/" + filename
    r_cable_path = r_file_path + "/" + filename
    # create directories if they do not exist
    tools.makeDir(cable_path)
    tools.makeDir(r_cable_path)
            
    #
    # get serial number of test boards
    #
    left_serialnumber = input(Fore.RED + "Enter SN of left board: " + Fore.GREEN)
    left_serialnumber = left_serialnumber.strip()
    while left_serialnumber == "" :
        left_serialnumber = input(Fore.RED + "Enter SN of left board: " + Fore.GREEN)
        left_serialnumber = left_serialnumber.strip()

    right_serialnumber = input(Fore.RED + "Enter SN of right board: " + Fore.GREEN)
    right_serialnumber = right_serialnumber.strip()
    while right_serialnumber == "" :
        right_serialnumber = input(Fore.RED + "Enter SN of right board: " + Fore.GREEN)
        right_serialnumber = right_serialnumber.strip()
        
    left_serialnumber = left_serialnumber.upper()
    right_serialnumber = right_serialnumber.upper()
        
    #
    # brief operator notes
    #
    operator_notes = input(Fore.RED + "Operator notes: " + Fore.GREEN)

    #
    # guarantee caps lock is off
    #
    if ctypes.WinDLL("User32.dll").GetKeyState(0x14) :
        print(Fore.RED + "CAPSLOCK is on - turning off." + Fore.RESET)
        pygui.press('capslock')

    keys = list(cable_mapping.keys())

    if RUN_4PT_DC_RES_CALIBRATION:
        print(Fore.GREEN + "")
        print("Running 4-point DC resistance calibration.")
    elif RUN_4PT_DC_RES or RUN_EYE_BERT_AREA:
        print(Fore.GREEN + "")
        if branch:
            print(f"Taking data for cable {cable}, branch {branch}.")
        else:
            print(f"Taking data for cable {cable}.")
    else:
        print(Fore.GREEN + "")
        print("Nothing to do (based on run flags)...")

    # 4-point DC resistance calibration
    if RUN_4PT_DC_RES_CALIBRATION:
        print(Fore.GREEN + "")
        print("--------------------------------------------")
        print("Beginning 4-point DC resistance calibration.")
        print("--------------------------------------------")

        # Note: Make sure to use a new calibration file name; the calibration file you specify will be overwritten!
        calibration_data = {}
        calibration_file = input(Fore.RED + f"Please enter a new calibration json file, including e-link type (4_point_DC_CalibrationRevB_Type3p2_v1.json): " + Fore.GREEN)
        print(f"Calibration data will be saved to {calibration_file}. This file will be overwritten.")
        
        # Confirm that user wants to continue
        valid_inputs = ["y", "n"]
        accept = input(Fore.RED + f"Do you want to proceed? (y/n): " + Fore.GREEN).lower()
        while accept not in valid_inputs:
            accept = input(f"The input '{accept}' is not vaild. Do you want to proceed (y/n): ").lower()
        if accept == "y":
            print("Proceeding with calibration. Please connect SMA through lines for each channel as instructed.")
        if accept == "n":
            print("Exiting...")
            print(Fore.RED + "Terminating code 3: exit based on user input.")
            sys.exit(3)
                
        print("Measured calibration values (ohms):")

        # Loop over channels
        for key in keys :
            # skip key if it is "name"
            if key == "name" :
                continue
            # otherwise, we assume that the key is the channel
            else:
                channel = str(key)

            # get TX and RX paths from cable mapping
            txpath = b"tx " + bytes(cable_mapping[key]['tx'], 'utf-8')
            rxpath = b"rx " + bytes(cable_mapping[key]['rx'], 'utf-8')
            
            # pause for user to connect SMA through lines (P to P and N to N) for channel
            print(Fore.RED + f"Please connect SMA through lines (P to P and N to N) for channel {key}: txpath = {txpath.decode()} and rxpath = {rxpath.decode()}." + Fore.GREEN)
            user_ready = input(Fore.RED + f"Press enter when ready. " + Fore.GREEN)

            # take measurements; do not subtract anything
            eb.connection(txpath+b"\r\n")
            eb.connection(rxpath+b"\r\n")
            eb.LED(2,"ON")
            eb.MODE(b"MODE DMM +\r\n")
            positive = round(dmm.reading(),2)
            eb.MODE(b"MODE DMM -\r\n")
            negative = round(dmm.reading(),2)

            # print results
            print(" - channel {0}: {1}_p = {2:.2f}, {3}_n = {4:.2f}".format(key, key, positive, key, negative))

            # Ask user to accept or redo measurement
            valid_inputs = ["y", "n"]
            accept = input(Fore.RED + f"Please review and accept (y) or redo (n) this measurement: " + Fore.GREEN).lower()
            while accept not in valid_inputs:
                accept = input(f"The input '{accept}' is not vaild. Please accept (y) or redo (n) this measurement: ").lower()
            
            while accept == "n":            
                user_ready = input(Fore.RED + f"Press enter when ready. " + Fore.GREEN)

                # take measurements; do not subtract anything
                eb.connection(txpath+b"\r\n")
                eb.connection(rxpath+b"\r\n")
                eb.LED(2,"ON")
                eb.MODE(b"MODE DMM +\r\n")
                positive = round(dmm.reading(),2)
                eb.MODE(b"MODE DMM -\r\n")
                negative = round(dmm.reading(),2)

                # print results
                print(" - channel {0}: {1}_p = {2:.2f}, {3}_n = {4:.2f}".format(key, key, positive, key, negative))

                # Ask user to accept or redo measurement
                valid_inputs = ["y", "n"]
                accept = input(Fore.RED + f"Please review and accept (y) or redo (n) this measurement: " + Fore.GREEN).lower()
                while accept not in valid_inputs:
                    accept = input(f"The input '{accept}' is not vaild. Please accept (y) or redo (n) this measurement: ").lower()

            if accept == "y":
                print("Proceeding with calibration.")
            
            # save calibration data
            calibration_data[key + "_p"] = positive
            calibration_data[key + "_n"] = negative

        # Save calibration data to json file
        print("")
        print(f"Saving calibration data to {calibration_file}.")
        with open(calibration_file, "w") as write_file:
            json.dump(calibration_data, write_file, indent=4)
        
        print("The calibration is complete!")
        print("Exiting...")
        sys.exit(4)

    # 4-point DC resistance measurements
    if RUN_4PT_DC_RES:
        print(Fore.GREEN + "")
        print("---------------------------------------------")
        print("Beginning 4-point DC resistance measurements.")
        print("---------------------------------------------")
        
        measurement_data = {}
        calibration_data = {}
        calibration_file = GetDCCalibrationFile(cable_type)

        print(f"Using the calibration file {calibration_file}.")
        
        # Load calibration data from json file
        with open(calibration_file, "r") as read_file:
            calibration_data = json.load(read_file)
        
        print("Measurements after subtracting calibration values (ohms):")

        # Loop over channels
        for key in keys :
            # skip key if it is "name"
            if key == "name" :
                continue
            # otherwise, we assume that the key is the channel
            else:
                channel = str(key)
            
            # get calibration data for channel (for both P and N lines)
            pos_path = calibration_data[key + "_p"]
            neg_path = calibration_data[key + "_n"]
            
            # get TX and RX paths from cable mapping
            txpath = b"tx " + bytes(cable_mapping[key]['tx'], 'utf-8')
            rxpath = b"rx " + bytes(cable_mapping[key]['rx'], 'utf-8')

            # take measurements and subtract calibration values
            eb.connection(txpath+b"\r\n")
            eb.connection(rxpath+b"\r\n")
            eb.LED(2,"ON")
            eb.MODE(b"MODE DMM +\r\n")
            positive = round(dmm.reading(),2) - pos_path
            eb.MODE(b"MODE DMM -\r\n")
            negative = round(dmm.reading(),2) - neg_path
            
            # key with p and n
            key_p = key + "_p"
            key_n = key + "_n"

            # print results
            tools.PrintDCValues(key, key_p, key_n, positive, negative)
            
            # save measurement data
            measurement_data[key_p] = positive
            measurement_data[key_n] = negative

        # get bad 4-point DC channels
        bad_channels = tools.GetBadDCChannels(measurement_data)
        # print bad 4-point DC channels
        tools.PrintBadDCChannels(bad_channels)

        # get date and time
        now = datetime.datetime.now()
        date_now = now.strftime("%Y-%m-%d")
        time_now = now.strftime("%H:%M:%S")

        # results for this cable (and branch, if applicable)
        results_for_cable = {
            "cable"         : cable,
            "date"          : date_now,
            "time"          : time_now,
            "operator"      : operator,
            "left_SN"       : left_serialnumber, 
            "right_SN"      : right_serialnumber,
            "notes"         : operator_notes
        }

        # Cable with branch
        if branch:
            results_for_cable["branch"] = branch
        
        # save data for all channels (p and n)
        for x in all_channels:
            results_for_cable[x] = measurement_data[x]
        
        # save results for this cable (and branch, if applicable)
        dc_resistance_results.update(results_for_cable)

        # update XLS file, create new entries as needed
        wb = Workbook()
        path = "R:/BEAN_GRP/EyeBertAutomation/"
                
        # Use different spreasheets for each cable type
        file_name = f"DCResistanceAutomationRevB_Type_{cable_type}.xlsx"

        full_file_path = path + file_name

        # check if file exists
        fileExists = os.path.exists(full_file_path)

        # if file exists, check if file is open
        if fileExists:
            keep_trying = True
            while keep_trying:
                try:
                    os.rename(full_file_path, full_file_path)
                    keep_trying = False
                except OSError:
                    print(Fore.RED + file_name + " summary file is open. Please close.")
                    x = input(Fore.RED + "Press ENTER when ready to retry. " + Fore.GREEN)
                    keep_trying = True
                
        # table headers: defines order of columns in table
        headers = ["cable"]

        # Cable with branch
        if branch:
            headers += ["branch"]

        headers += ["date", "time"]
        headers += all_channels
        headers += ["operator", "left_SN", "right_SN", "notes"]
        
        # if file does not exist, create file with table headers
        if not fileExists:
            # create file
            print(Fore.LIGHTRED_EX + "\t" + file_name + " summary file does not exist. Creating file.")
            ws = wb.active            
            ws.append(headers)
            col = get_column_letter(1)
            ws.column_dimensions[col].bestFit = True
            wb.save(full_file_path)
        else:
            # load existing copy
            print(Fore.GREEN + "Opening summary file " + file_name)
            wb = load_workbook(filename = full_file_path)
            ws = wb.active
        
        print(Fore.GREEN + "Adding data...")
                
        newdata = [dc_resistance_results[x] for x in headers]

        ws.append(newdata)
        col = get_column_letter(1)
        ws.column_dimensions[col].bestFit = True
        wb.save(full_file_path)

        print(Fore.GREEN + Style.BRIGHT + "4-point DC resistance measuremsnts are complete!" + Fore.RESET + Style.RESET_ALL)

    # Eye BERT area measurements
    if RUN_EYE_BERT_AREA:
        print(Fore.GREEN + "")
        print("-------------------------------------")
        print("Beginning Eye BERT area measurements.")
        print("-------------------------------------")

        eb.MODE(b"MODE KC705\r\n")

        # Loop over channels
        for key in keys :
            # skip key if it is "name"
            if key == "name" :
                continue
            # otherwise, we assume that the key is the channel
            else:
                channel = str(key)
            
            # base test name:
            # - cable has branches: cable_branch_channel    
            # - cable does not have branches: cable_channel
            if branch:
                temp_name = f"{cable}_{branch}_{channel}"
            else:
                temp_name = f"{cable}_{channel}"
            # replace spaces to get final base test name:
            test_name = temp_name.replace(" ", "_")

            txpath = b"tx " + bytes(cable_mapping[key]['tx'], 'utf-8')
            rxpath = b"rx " + bytes(cable_mapping[key]['rx'], 'utf-8')
            
            print(Fore.GREEN + Style.BRIGHT + f"{test_name}" + Style.NORMAL)
            print(Fore.GREEN + f"\ttxpath = {txpath.decode()}")
            print(Fore.GREEN + f"\trxpath = {rxpath.decode()}")

            eb.connection(txpath+b"\r\n")
            eb.connection(rxpath+b"\r\n")
            eb.LED(2,"ON")

            # if old temp.csv exists, erase
            if os.path.exists(file_path+"/temp.csv") :
                os.remove(file_path+"/temp.csv")

            # makes assumption that EyeBERT KC704 (Vivado) is already running
            # with bitstream loaded
            # guarantee it is maximized and has focus
            try :
                pygui.getWindowsWithTitle("Vivado 2020.2")[0].maximize()
                time.sleep(0.5)
                pygui.getWindowsWithTitle("Vivado 2020.2")[0].activate()
                time.sleep(0.5)
                # send F5 to reset layout of Vivado
                pygui.press('f5')
            except :
                print(Fore.RED + "Vivado 2020.2 and EyeBert bitstream not running")
                print(Fore.RED + "Terminating code 2" + Fore.RESET)
                sys.exit(2)

            # create custom "source eye_and_save.tcl" with the cable name & path included
            eye_and_save = "C:/Users/Public/Documents/cable_tests/eye_and_save.tcl"
            with open (eye_and_save, 'w') as f :
                f.write("source create_scan_0_RevB.tcl\r\n")
                f.write(f"set_property DESCRIPTION {test_name} [get_hw_sio_scans SCAN_0]\r\n")
                f.write("run_hw_sio_scan [lindex [get_hw_sio_scans {SCAN_0}] 0]\r\n")
                f.write("wait_on_hw_sio_scan [lindex [get_hw_sio_scans {SCAN_0}] 0]\r\n")
                f.write('write_hw_sio_scan -force "C:/Users/Public/Documents/automation_results/temp.csv" [get_hw_sio_scans {SCAN_0}]\r\n')
            
            # pygui hotkeys: https://pyautogui.readthedocs.io/en/latest/keyboard.html
            # Hotkey to select Tcl Console and put the cursor in the command box
            print(Fore.GREEN + "Pressing hotkey for Tcl Console!")
            time.sleep(0.25)
            pygui.hotkey('ctrl', 'shift', 't')
            time.sleep(0.25)

            # launch the test via TCL script
            #rwy polarity TCL goes here before the scan: it is a link setting, not a scan setting
            pygui.write("source eye_and_save.tcl\n", interval = 0.01)

            # wait a bit
            print(Fore.GREEN + "Pausing to allow EyeBERT to complete...")
            time.sleep(15)

            # loop looking for temp.csv
            print(Fore.GREEN + "Waiting for data file from Vivado...")
            time_to_wait = 20
            time_counter = 0
            while not os.path.exists(file_path+"/temp.csv") :
                time.sleep(1)
                time_counter += 1
                if time_counter > time_to_wait : break

            eb.LED(2,"OFF")

            # get data
            print(Fore.GREEN + "Parsing CSV file for test results...")
            open_area = -999.9
            with open(file_path+"/temp.csv", 'r') as file :
                csvreader = csv.reader(file, delimiter=',')
                # find the Open Area
                for row in csvreader :
                    if row[0] == "Open Area" :
                        open_area = int(row[1])
                        break

            with open(file_path+"/temp.csv", 'r') as file :
                search = list(csv.reader(file)) # convert to a list we can index through
                # find the top of eye
                top_of_eye = -999.9 # nonsense values
                # use column AH, start at row 22
                for r in range(21, 45, 1) :
                    value = float(search[r][33])
                    if value < 2.0e-7 : 
                        # we have found it
                        top_of_eye = float(search[r-1][0])
                        break
                # find the bottom of eye
                bottom_of_eye = 999.9
                for r in range(44, 20, -1) :
                    value = float(search[r][33])
                    if value < 2.0e-7 :
                        # we found it!
                        bottom_of_eye = float(search[r+1][0])
                        break

            print(Fore.GREEN + Style.BRIGHT + f" - Open area = {open_area}")
            print(Fore.GREEN + f" - top_of_eye = {top_of_eye}")
            print(Fore.GREEN + f" - bottom_of_eye = {bottom_of_eye}" + Style.NORMAL)

            src = file_path+"/temp.csv"
            dest = cable_path + "/" + test_name + ".csv"
            # check if destionation already exists
            counter = 2
            while os.path.exists(dest) == True :
                dest = cable_path + "/" + test_name + "_" + str(counter) + ".csv"
                counter = counter + 1

            os.rename(src,dest)

            # Copy data csv file to R drive
            copyfilename = dest.replace(file_path+"/","")
            newdest = r_file_path + "/" + copyfilename
            if verbose:
                print(f"data path (1): {dest}")
                print(f"data path (2): {newdest}")
            shutil.copyfile(dest, newdest)
            
            # EyeBERT Template Analysis

            reference_template_file = "reference_template_RevB_v3.csv"
            print(f"Using this reference template data file: {reference_template_file}")
            
            # Create EyeBERTFile object to read data from file
            eyebert = EyeBERTFile(cable, branch, channel, file_path)
            
            # Call analyze method to obtain graphs and properties
            analysis = eyebert.analyze()

            # Warning: .getPath() must be called AFTER .analyze()
            refPath = eyebert.getPath()
            
            if verbose:
                print(f"refPath = {refPath}")

            analysis.setPath(refPath)

            if analysis.verify(): # Only continue if template passes verifcation

                template = analysis.createTemplate()

                ref = Reference("540", branch, "CMD", reference_template_file, refPath)
                refTemp = ref.createTemplate()
                
                # EDIT: reference needs to be a template object
                template.plot(refTemp) 
                print("\n")
                
                # Get template analysis results
                num_zeros   = template.getZeros()
                num_ones    = template.getOnes()
                out_points  = template.getOutCounts()
                in_points   = template.getInCounts()
            else:
                print(f"Error for cable {cable}, branch {branch}, and channel {channel.upper()}: Template failed verification step.")

            # Make the screen capture
            print(Fore.GREEN + "Creating screen capture...")
            pygui.moveTo(355,800)
            pygui.click()
            screenshot = cable_path + "/" + test_name + ".png"
            counter = 2
            while os.path.exists(screenshot) == True :
                screenshot = cable_path + "/" + test_name + "_" + str(counter) + ".png"
                counter = counter + 1

            # Save original screenshot 
            pygui.screenshot(screenshot,(360,190,1300,540))

            # Copy screenshot to the R drive
            copyfilename = screenshot.replace(file_path+"/","")
            newdest = r_file_path + "/" + copyfilename
            if verbose:
                print(f"screenshot path (1): {screenshot}")
                print(f"screenshot path (2): {newdest}")
            shutil.copyfile(screenshot, newdest)

            # Copy template plot to the R drive
            oldTemplatePath = refPath + "template_plots.pdf"
            endPath = oldTemplatePath.split("/automation_results/")[-1]
            newTemplatePath = r_file_path + "/" + endPath 
            if verbose:
                print(f"template plot path (1): {oldTemplatePath}")
                print(f"template plot path (2): {newTemplatePath}")
            shutil.copyfile(oldTemplatePath, newTemplatePath)
            
            # get date and time
            now = datetime.datetime.now()
            date_now = now.strftime("%Y-%m-%d")
            time_now = now.strftime("%H:%M:%S")
                
            # results for this channel
            results_for_channel = {
                key : {
                        "cable"         : cable,
                        "channel"       : channel,
                        "date"          : date_now,
                        "time"          : time_now,
                        "open_area"     : open_area, 
                        "top_eye"       : top_of_eye, 
                        "bottom_eye"    : bottom_of_eye,
                        "num_zeros"     : num_zeros,
                        "num_ones"      : num_ones,
                        "out_points"    : out_points,
                        "in_points"     : in_points,
                        "operator"      : operator,
                        "left_SN"       : left_serialnumber, 
                        "right_SN"      : right_serialnumber,
                        "notes"         : operator_notes
                }
            }
            
            # Cable with branch
            if branch:
                results_for_channel[key]["branch"] = branch
            
            # save results for this channel
            eye_bert_results.update(results_for_channel)

        #end keys loop

        # update XLS file, create new entries as needed
        wb = Workbook()
        path = "R:/BEAN_GRP/EyeBertAutomation/"

        # Use different spreasheets for each cable type
        file_name = f"EyeBERTautomationRevB_Type_{cable_type}.xlsx"
        
        full_file_path = path + file_name

        # check if file exists
        fileExists = os.path.exists(full_file_path)

        # if file exists, check if file is open
        if fileExists:
            keep_trying = True
            while keep_trying:
                try:
                    os.rename(full_file_path, full_file_path)
                    keep_trying = False
                except OSError:
                    pygui.getWindowsWithTitle("Vivado 2020.2")[0].minimize()
                    print(Fore.RED + file_name + " summary file is open. Please close.")
                    x = input(Fore.RED + "Press ENTER when ready to retry. " + Fore.GREEN)
                    keep_trying = True
                
        # table headers: defines order of columns in table
        headers = ["cable"]
        
        # Cable with branch
        if branch:
            headers += ["branch"]
        
        headers += ["channel", "date", "time", "open_area", "top_eye", "bottom_eye", "num_zeros", "num_ones", "out_points", "in_points",
                    "operator", "left_SN", "right_SN", "notes"]
        
        # if file does not exist, create file with table headers
        if not fileExists:
            # create file
            print(Fore.LIGHTRED_EX + "\t" + file_name + " summary file does not exist. Creating file.")
            ws = wb.active            
            ws.append(headers)
            col = get_column_letter(1)
            ws.column_dimensions[col].bestFit = True
            wb.save(full_file_path)
        else:
            # load existing copy
            print(Fore.GREEN + "Opening summary file " + file_name)
            wb = load_workbook(filename = full_file_path)
            ws = wb.active

        print(Fore.GREEN + "Adding data...")
        keys = list(eye_bert_results)
        for key in keys:
            newdata = [eye_bert_results[key][x] for x in headers]
            ws.append(newdata)
        col = get_column_letter(1)
        ws.column_dimensions[col].bestFit = True
        wb.save(full_file_path)

        pygui.getWindowsWithTitle("Vivado 2020.2")[0].minimize()
        print(Fore.GREEN + Style.BRIGHT + "Eye BERT area measuremsnts are complete!" + Fore.RESET + Style.RESET_ALL)

        #excel_start_return = os.system('start "excel" ' + full_file_path)

if __name__ == "__main__":
    main()
